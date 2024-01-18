import openpyxl
from openpyxl.utils import get_column_letter
import argparse
import traceback


def get_excel_coordinates(row, column):
    '''
    row,col から excel 座標を返す
    ex: (1, 1) = A1
    '''
    return get_column_letter(column) + str(row)


def get_borders(cell):
    '''
    セルの罫線の状態を確認する
    '''
    borders = []
    if cell.border.left.style:
        borders.append('LEFT')
    if cell.border.right.style:
        borders.append('RIGHT')
    if cell.border.top.style:
        borders.append('TOP')
    if cell.border.bottom.style:
        borders.append('BOTTOM')
    return borders


def get_bgcolor(cell):
    '''
    セルの背景色を取得
    '''
    bgcolor = cell.fill.start_color.index
    return bgcolor


def get_merged_cells(sheet, cell):
    '''
    セルのマージ情報を取得
    '''
    cell_index = cell.coordinate
    # シート内の全マージリストからcellに該当するものがあるか検索
    for range_ in sheet.merged_cells.ranges:
        # sheet.merged_cells.ranges = {<MergedCellRange B2:C2>, <MergedCellRange B3:C3>, ...}
        merged_cells = list(openpyxl.utils.rows_from_range(str(range_)))
        for row in merged_cells:  # (A1,A2,A3), (B1,B2,B3), ... セルのマージ座標
            if cell_index in row:
                # マージあり
                return merged_cells
    # マージなし
    return None


def main(args):
    # excel 座標系変換実験
    wb = openpyxl.load_workbook(args.xlsfile)
    print("wb.worksheets:{}".format(wb.worksheets))  # worksheetオブジェクトのリストを取得
    print("wb.sheetnames:{}".format(wb.sheetnames))  # シート名のリストを取得
    ws = wb["オフィス"]
    print("ws['A1']:{}".format(ws['A1'].value))  # A1 の内容を取得
    for index in range(100):  # とりあえず１００row検索することにする
        row = index + 2  # row は１始まりでかつ、A1 にはすでに記載があるので +2
        if not ws.cell(row, 1).value:  # Noneの場合、番号なしなので終端とする
            break
        # 以下で、番号,name-jp,name のそれぞれのセル情報を出力する
        #  - value: セルの値
        #  - borders: 罫線情報
        #  - merged_cells: マージしている場合、マージしているセルリスト
        # 番号のセル情報を取得
        cell = ws.cell(row, 1)
        excel_pos = get_excel_coordinates(row, 1)
        bgcolor = get_bgcolor(cell)
        borders = get_borders(cell)
        merged_cells = get_merged_cells(ws, cell)
        print("number {}:value={},bgcolor={},borders={},merged_cells={}".format(excel_pos, cell.value, bgcolor, ",".join(borders), merged_cells))
        # name-jpのセル情報を取得
        cell = ws.cell(row, 2)
        excel_pos = get_excel_coordinates(row, 2)
        bgcolor = get_bgcolor(cell)
        borders = get_borders(cell)
        merged_cells = get_merged_cells(ws, cell)
        print("name-jp {}:value={},bgcolor={},borders={},merged_cells={}".format(excel_pos, cell.value, bgcolor, ",".join(borders), merged_cells))
        # nameのセル情報を取得
        cell = ws.cell(row, 4)
        excel_pos = get_excel_coordinates(row, 4)
        bgcolor = get_bgcolor(cell)
        borders = get_borders(cell)
        merged_cells = get_merged_cells(ws, cell)
        print("name {}:value={},bgcolor={},borders={},merged_cells={}".format(excel_pos, cell.value, bgcolor, ",".join(borders), merged_cells))

    wb.close()  # book 閉じる
    print("complete")


if __name__ == '__main__':  # プログラム実行ポイント
    try:
        # 入力引数設定
        parser = argparse.ArgumentParser(description='xls読み込み')
        parser.add_argument('xlsfile', help='xls ファイル')  # 入力引数１：excelファイルであることを示す
        args = parser.parse_args()  # 入力引数取得
        main(args)
    except Exception:  # as e: # main() で発生する異常はすべてキャッチする
        t = traceback.format_exc()
        print("ERROR: {}".format(t))
