import openpyxl
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
import argparse
import traceback


def get_excel_coordinates(row, column):
    '''
    row,col から excel 座標を返す
    ex: (1, 1) = A1
    '''
    return get_column_letter(column) + str(row)


def get_row_col(excel_pos):
    '''
    excel 座標からrow,col位置を返す
    ex: A1 = (1, 1)
    '''
    col_letter, row = coordinate_from_string(excel_pos)  # 位置取得
    col = column_index_from_string(col_letter)
    return row, col


def create_border(top=False, bottom=False, left=False, right=False):
    '''
    セルに罫線を引く
    '''
    side = Side(style='thin', color='000000')  # 罫線のスタイルと色を設定
    border = Border(
        top=side if top else None,
        bottom=side if bottom else None,
        left=side if left else None,
        right=side if right else None
    )  # 罫線を設定
    return border


def main(args):
    # excel 座標系変換実験
    print("(2, 2)={}".format(get_excel_coordinates(2, 2)))  # row=2, col=2 をexcel座標系に変換すると B2
    print("(B2)={}".format(get_row_col('B2')))  # B2 を row, col に変換すると (2, 2)
    wb = openpyxl.Workbook()  # book 新規作成(勝手にゼロ番目のシートにシート名「Sheet」が作成されるので注意)
    wb.remove(wb["Sheet"])  # 勝手作成されたSheetを削除
    ws = wb.create_sheet(title='オフィス')
    # 本スクリプトは utf8-lf のため、文字コードを意識してexcelに代入する必要はない
    items = [
        {'name': 'word', 'name-jp': 'ワード'},
        {'name': 'excel', 'name-jp': 'エクセル'},
        {'name': 'powerpoint', 'name-jp': 'パワポ'},
        {'name': 'pdf', 'name-jp': 'ピーディーエフ'},
    ]
    ws['A1'].value = '書き込みサンプル'  # excel 座標系 A1 に '書き込みサンプル' を書き込み
    for index in range(len(items)):
        item = items[index]  # 書き込むitem情報
        border = create_border(top=True, bottom=True, left=True, right=True)  # ４方向すべて罫線を引く
        bg_gray = PatternFill(patternType="solid", fgColor="a0a0a0")  # 塗りつぶし色グレー
        # 以下の書き込みはrow,col形式で座標を設定している。row,colが (1, 1) の場合、 excel座標系は A1 となる
        row = index + 2  # row は１始まりでかつ、A1 にはすでに記載があるので +2
        # 番号を書き込み
        cell = ws.cell(row, 1)
        cell.value = str(index + 1)
        cell.border = border
        cell.fill = bg_gray  # 背景色をグレーにする
        # name-jp を書き込み
        cell = ws.cell(row, 1+1)
        cell.value = str(item['name-jp'])
        cell.border = border  # 罫線設定
        ws.merge_cells(start_row=row, start_column=1+1,  # セルマージ：開始row,col
                       end_row=row, end_column=1+2)  # セルマージ：終了row,col
        # name を書き込み
        cell = ws.cell(row, 1+3)
        cell.value = str(item['name'])
        cell.border = border  # 罫線設定
    wb.save(args.xlsfile)  # excel ファイル書き出し（入力引数１番）
    wb.close()  # book 閉じる
    print("complete")


if __name__ == '__main__':  # プログラム実行ポイント
    try:
        # 入力引数設定
        parser = argparse.ArgumentParser(description='xls読み込み')
        parser.add_argument('xlsfile', help='xls ファイル')  # 入力引数１：excelファイルであることを示す
        args = parser.parse_args()  # 入力引数取得
        main(args)
    except Exception:  # as e: # main() で発生する異常はすべてキャッチする
        t = traceback.format_exc()
        print("ERROR: {}".format(t))
